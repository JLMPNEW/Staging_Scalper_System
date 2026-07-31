from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
import logging
import multiprocessing
import re
import time
from typing import Any, Literal
from xml.etree import ElementTree as ET
from zipfile import BadZipFile, ZipFile


MACHINERY_FOOTNOTE_REPORT_KEYWORDS = (
    "backlog",
    "booking",
    "contract",
    "customer",
    "inventory",
    "order",
    "performance obligation",
    "remaining performance",
    "revenue",
    "segment",
)


@dataclass(frozen=True)
class DocumentText:
    text: str
    extraction_method: str
    page_count: int = 0
    ocr_used: bool = False
    warning: str = ""


def filing_summary_document_name(index_payload: dict[str, Any]) -> str:
    raw_items = (index_payload.get("directory") or {}).get("item") or []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        if name.lower() == "filingsummary.xml":
            return name
    return ""


def filing_summary_report_documents(
    filing_summary_xml: str,
    *,
    keywords: tuple[str, ...] = MACHINERY_FOOTNOTE_REPORT_KEYWORDS,
) -> set[str]:
    try:
        root = ET.fromstring(filing_summary_xml)
    except ET.ParseError:
        return set()
    normalized_keywords = tuple(re.sub(r"\s+", " ", keyword.strip().lower()) for keyword in keywords if keyword.strip())
    selected: set[str] = set()
    for report in root.iter():
        if report.tag.rsplit("}", 1)[-1].lower() != "report":
            continue
        fields = {child.tag.rsplit("}", 1)[-1].lower(): str(child.text or "").strip() for child in report}
        document_name = fields.get("htmlfilename", "")
        if not document_name.lower().endswith((".htm", ".html", ".xhtml")):
            continue
        description = " ".join(
            fields.get(field, "") for field in ("shortname", "longname", "menucategory", "role")
        ).lower()
        description = re.sub(r"\s+", " ", description)
        if any(keyword in description for keyword in normalized_keywords):
            selected.add(document_name)
    return selected


def decode_text_document(payload: bytes) -> DocumentText:
    if payload.startswith((b"\xff\xfe", b"\xfe\xff")):
        return DocumentText(payload.decode("utf-16"), "text_decode_utf16")
    if payload.startswith(b"\xef\xbb\xbf"):
        return DocumentText(payload.decode("utf-8-sig"), "text_decode_utf8")
    try:
        return DocumentText(payload.decode("utf-8"), "text_decode_utf8")
    except UnicodeDecodeError:
        # Windows-1252 filings (smart quotes, en-dashes, NBSP, currency
        # symbols) are common; lossy U+FFFD replacement breaks date regexes
        # ("March 31,<?>2026") and unit detection. cp1252 decodes any byte, so
        # this never raises.
        return DocumentText(payload.decode("cp1252"), "text_decode_cp1252")


def _pypdf_text_unbounded(payload: bytes, *, max_pages: int) -> DocumentText:
    try:
        from pypdf import PdfReader  # type: ignore[import-not-found]
    except ImportError:
        return DocumentText("", "pdf_unavailable", warning="pypdf_not_installed")

    try:
        logging.getLogger("pypdf").setLevel(logging.ERROR)
        reader = PdfReader(BytesIO(payload), strict=False)
        page_count = len(reader.pages)
        pages = reader.pages[:max_pages] if max_pages > 0 else reader.pages
        text = "\n\n".join(str(page.extract_text() or "") for page in pages).strip()
    except Exception as exc:  # pragma: no cover - library-specific malformed PDF failures
        return DocumentText(
            "",
            "pdf_pypdf_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    return DocumentText(text, "pdf_pypdf", page_count=page_count)


def _pymupdf_text_unbounded(payload: bytes, *, max_pages: int) -> DocumentText:
    try:
        import pymupdf  # type: ignore[import-not-found]
    except ImportError:
        return DocumentText(
            "",
            "pdf_pymupdf_unavailable",
            warning="pymupdf_not_installed",
        )
    pages: list[str] = []
    try:
        with pymupdf.open(stream=payload, filetype="pdf") as document:
            page_count = document.page_count
            limit = min(page_count, max_pages) if max_pages > 0 else page_count
            for page_number in range(limit):
                pages.append(document.load_page(page_number).get_text("text"))
    except Exception as exc:
        return DocumentText(
            "",
            "pdf_pymupdf_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    return DocumentText(
        "\n\n".join(pages).strip(),
        "pdf_pymupdf_targeted_recovery",
        page_count=page_count,
    )


def _ocr_pdf_text_unbounded(payload: bytes, *, max_pages: int) -> DocumentText:
    try:
        import pymupdf  # type: ignore[import-not-found]
        import pytesseract  # type: ignore[import-not-found]
        from PIL import Image  # type: ignore[import-not-found]
    except ImportError:
        return DocumentText("", "pdf_ocr_unavailable", warning="pdf_ocr_dependencies_not_installed")

    pages: list[str] = []
    try:
        with pymupdf.open(stream=payload, filetype="pdf") as document:
            page_count = document.page_count
            limit = min(page_count, max_pages) if max_pages > 0 else page_count
            for page_number in range(limit):
                page = document.load_page(page_number)
                image_bytes = page.get_pixmap(matrix=pymupdf.Matrix(2.0, 2.0), alpha=False).tobytes("png")
                with Image.open(BytesIO(image_bytes)) as image:
                    pages.append(str(pytesseract.image_to_string(image) or ""))
    except Exception as exc:  # pragma: no cover - external OCR/PDF runtime failures
        return DocumentText(
            "",
            "pdf_ocr_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    return DocumentText(
        "\n\n".join(pages).strip(),
        "pdf_ocr",
        page_count=page_count,
        ocr_used=True,
    )


def _pdf_worker(
    method: Literal["pypdf", "pymupdf", "ocr"],
    payload: bytes,
    max_pages: int,
    sender: Any,
) -> None:
    try:
        if method == "pypdf":
            result = _pypdf_text_unbounded(payload, max_pages=max_pages)
        elif method == "pymupdf":
            result = _pymupdf_text_unbounded(payload, max_pages=max_pages)
        else:
            result = _ocr_pdf_text_unbounded(payload, max_pages=max_pages)
        sender.send(result)
    except BaseException as exc:  # pragma: no cover - protects the parent from worker failures
        sender.send(
            DocumentText(
                "",
                f"pdf_{method}_worker_failed",
                warning=f"{type(exc).__name__}:{exc}",
            )
        )
    finally:
        sender.close()


def _bounded_pdf_extract(
    method: Literal["pypdf", "pymupdf", "ocr"],
    payload: bytes,
    *,
    max_pages: int,
    timeout_sec: float,
) -> DocumentText:
    if timeout_sec <= 0:
        if method == "pypdf":
            return _pypdf_text_unbounded(payload, max_pages=max_pages)
        if method == "pymupdf":
            return _pymupdf_text_unbounded(payload, max_pages=max_pages)
        return _ocr_pdf_text_unbounded(payload, max_pages=max_pages)

    context = multiprocessing.get_context("spawn")
    receiver, sender = context.Pipe(duplex=False)
    process = context.Process(
        target=_pdf_worker,
        args=(method, payload, max_pages, sender),
        daemon=True,
        name=f"machinery-pdf-{method}",
    )
    try:
        process.start()
        sender.close()
    except Exception as exc:
        receiver.close()
        sender.close()
        return DocumentText(
            "",
            f"pdf_{method}_worker_failed",
            warning=f"worker_start_failed:{type(exc).__name__}:{exc}",
        )

    result: DocumentText | None = None
    deadline = time.monotonic() + timeout_sec
    try:
        while time.monotonic() < deadline:
            remaining = max(deadline - time.monotonic(), 0.0)
            if receiver.poll(min(0.1, remaining)):
                received = receiver.recv()
                if isinstance(received, DocumentText):
                    result = received
                break
            if not process.is_alive():
                break
    except (EOFError, OSError) as exc:
        result = DocumentText(
            "",
            f"pdf_{method}_worker_failed",
            warning=f"worker_pipe_failed:{type(exc).__name__}:{exc}",
        )
    finally:
        receiver.close()

    timed_out = process.is_alive()
    if timed_out:
        process.terminate()
        process.join(timeout=5.0)
    else:
        process.join(timeout=1.0)
    if process.is_alive():  # pragma: no cover - terminate is reliable on supported runtimes
        process.kill()
        process.join(timeout=1.0)

    if result is not None:
        return result
    if timed_out:
        return DocumentText(
            "",
            f"pdf_{method}_timeout",
            warning=f"extraction_timeout_seconds:{timeout_sec:g}",
        )
    if process.exitcode not in (0, None):
        return DocumentText(
            "",
            f"pdf_{method}_worker_failed",
            warning=f"worker_exitcode:{process.exitcode}",
        )
    return DocumentText("", f"pdf_{method}_worker_failed", warning="worker_returned_no_result")


def extract_pdf_text_pymupdf(
    payload: bytes,
    *,
    max_pages: int = 250,
    extraction_timeout_sec: float = 30.0,
) -> DocumentText:
    """Targeted local fallback for already-dispositioned PDF limitations."""
    return _bounded_pdf_extract(
        "pymupdf",
        payload,
        max_pages=max_pages,
        timeout_sec=extraction_timeout_sec,
    )


def extract_pdf_text(
    payload: bytes,
    *,
    enable_ocr: bool,
    max_pages: int = 250,
    minimum_text_characters: int = 200,
    extraction_timeout_sec: float = 30.0,
) -> DocumentText:
    direct = _bounded_pdf_extract(
        "pypdf",
        payload,
        max_pages=max_pages,
        timeout_sec=extraction_timeout_sec,
    )
    if len(direct.text.strip()) >= minimum_text_characters or not enable_ocr:
        if not enable_ocr and len(direct.text.strip()) < minimum_text_characters and not direct.warning:
            # An image-only PDF "succeeds" in pypdf with empty text. Without a
            # warning the adapter emits no evidence at all and the document is
            # mislabeled as issuer non-disclosure instead of PARSER_FAILURE.
            return DocumentText(
                direct.text,
                direct.extraction_method,
                page_count=direct.page_count,
                warning="pdf_no_native_text_ocr_disabled",
            )
        return direct
    ocr = _bounded_pdf_extract(
        "ocr",
        payload,
        max_pages=max_pages,
        timeout_sec=extraction_timeout_sec,
    )
    if len(ocr.text.strip()) > len(direct.text.strip()):
        return ocr
    if direct.warning and ocr.warning:
        return DocumentText(
            direct.text,
            direct.extraction_method,
            page_count=direct.page_count,
            warning=f"{direct.warning};{ocr.warning}",
        )
    return direct


def extract_xlsx_text(payload: bytes) -> DocumentText:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError:
        return DocumentText(
            "",
            "xlsx_unavailable",
            warning="openpyxl_not_installed",
        )
    try:
        workbook = load_workbook(
            BytesIO(payload),
            read_only=True,
            data_only=True,
        )
        lines: list[str] = []
        for worksheet in workbook.worksheets:
            lines.append(f"[Worksheet: {worksheet.title}]")
            for row in worksheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    lines.append("\t".join(values))
        workbook.close()
    except Exception as exc:
        return DocumentText(
            "",
            "xlsx_openpyxl_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    return DocumentText("\n".join(lines), "xlsx_openpyxl")


def extract_docx_text(payload: bytes) -> DocumentText:
    try:
        with ZipFile(BytesIO(payload)) as archive:
            document_xml = archive.read("word/document.xml")
    except (BadZipFile, KeyError, OSError) as exc:
        return DocumentText(
            "",
            "docx_zip_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    try:
        root = ET.fromstring(document_xml)
    except ET.ParseError as exc:
        return DocumentText(
            "",
            "docx_xml_failed",
            warning=f"{type(exc).__name__}:{exc}",
        )
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if paragraph.tag.rsplit("}", 1)[-1].lower() != "p":
            continue
        text = " ".join(
            str(node.text or "").strip()
            for node in paragraph.iter()
            if node.tag.rsplit("}", 1)[-1].lower() == "t" and str(node.text or "").strip()
        )
        if text:
            paragraphs.append(text)
    return DocumentText("\n".join(paragraphs), "docx_xml")


def extract_document_text(
    payload: bytes,
    *,
    document_name: str,
    content_type: str = "",
    enable_pdf_ocr: bool = False,
    max_pdf_pages: int = 250,
    max_pdf_bytes: int = 25_000_000,
    pdf_extraction_timeout_sec: float = 30.0,
) -> DocumentText:
    suffix = document_name.lower().rsplit(".", 1)[-1] if "." in document_name else ""
    normalized_content_type = content_type.lower()
    if suffix == "xlsx" or "spreadsheetml" in normalized_content_type:
        return extract_xlsx_text(payload)
    if suffix == "docx" or (payload.startswith(b"PK") and "msword" in normalized_content_type):
        return extract_docx_text(payload)
    if suffix == "doc" and payload.startswith(b"\xd0\xcf\x11\xe0"):
        return DocumentText(
            "",
            "legacy_word_binary_requires_conversion",
            warning="legacy_word_binary_requires_offline_conversion",
        )
    if suffix == "pdf" or "application/pdf" in content_type.lower() or payload.startswith(b"%PDF-"):
        if max_pdf_bytes > 0 and len(payload) > max_pdf_bytes:
            return DocumentText(
                "",
                "pdf_size_limit",
                warning=f"document_bytes:{len(payload)} exceeds max_pdf_bytes:{max_pdf_bytes}",
            )
        return extract_pdf_text(
            payload,
            enable_ocr=enable_pdf_ocr,
            max_pages=max_pdf_pages,
            extraction_timeout_sec=pdf_extraction_timeout_sec,
        )
    return decode_text_document(payload)
