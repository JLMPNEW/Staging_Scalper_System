#!/usr/bin/env python3
"""Backfill TRUE first-print CFNAI / CFNAI-MA3 vintages that ALFRED lacks.

ALFRED's CFNAI archive begins with vintage 2011-05-23; every earlier monthly
release only survives as the (already-revised) values carried inside that first
2011-05-23 edition. This importer recovers the genuine *as-published* first print
for each reference month by reading the Chicago Fed's own "CFNAI Historical
(Real-time) Data" spreadsheets, in which every release edition is preserved as a
frozen column keyed by its headline reference month.

Source (public, no key):
    Directory endpoint   https://data.chicagofed.org/cfed-drm-chicago/CFNAI
    Real-time workbooks  https://api.data.chicagofed.org/CFNAI/cfnai-realtime-{1,2,3}-xlsx.xlsx

Layout of each workbook's ``cfnai_realtime`` sheet:
    * column A         = reference month (first-of-month datetime)
    * header row 0     = edition keys ``CF{month}{year}`` (CFNAI) and
                         ``CF3{month}{year}`` (CFNAI-MA3), where month/year is the
                         *headline* (newest) reference month of that release edition
    * ``release_dates`` sheet maps each ``cfnai_month`` -> actual ``release_date``.

The first print of reference month M is therefore the value in the edition whose
headline month is M, i.e. matrix[row=M][col=CF{M.month}{M.year}], published on the
release_date that ``release_dates`` records for cfnai_month == M.

Rows are written to ``macro_observation_raw`` with the exact registry identity of
the existing ALFRED CFNAI rows (source_name=fred_alfred, source_series_id CFNAI /
CFNAIMA3), release_date = vintage_date = the true historical publication date, and
the standard dedupe key, so they merge with (never duplicate) ALFRED editions.

Default is a DRY RUN (parses + prints, writes nothing). Pass --apply to write.
"""
from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from macro_raw_config import (
    configure_pipeline_logging,
    connect_sqlite,
    load_macro_raw_config,
    resolve_config_path,
    resolve_db_path,
    resolve_path,
    utc_now_iso,
)
from macro_registry import MetricSpec, load_metric_registry
from macro_storage import (
    _insert_artifacts,
    _upsert_observations,
    finish_run,
    init_db,
    start_run,
)
from macro_types import ObservationRecord, SourceArtifact

logger = logging.getLogger(__name__)

DIRECTORY_ENDPOINT = "https://data.chicagofed.org/cfed-drm-chicago/CFNAI"
FALLBACK_REALTIME_URLS = [
    "https://api.data.chicagofed.org/CFNAI/cfnai-realtime-1-xlsx.xlsx",
    "https://api.data.chicagofed.org/CFNAI/cfnai-realtime-2-xlsx.xlsx",
    "https://api.data.chicagofed.org/CFNAI/cfnai-realtime-3-xlsx.xlsx",
]
BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
)

CFNAI_REGISTRY_KEY = "us_cfnai"
CFNAI_MA3_REGISTRY_KEY = "us_cfnai_ma3"
VALUE_ROUND_DP = 2

# Gap window keyed by reference (observation) month. ALFRED's first CFNAI vintage
# is 2011-05-23 (headline reference month 2011-04); everything strictly earlier is
# unrecoverable from ALFRED. Defaults span the spec-frozen fill window.
DEFAULT_START_MONTH = "2007-06"
DEFAULT_END_MONTH = "2011-04"


@dataclass(frozen=True)
class FirstPrint:
    ref_year: int
    ref_month: int
    release_date: date
    cfnai: float
    cfnai_ma3: float
    source_url: str


def _http_get(url: str, *, data: bytes | None = None) -> bytes:
    headers = {"User-Agent": BROWSER_UA, "Accept": "*/*"}
    req = urllib.request.Request(url, data=data, headers=headers)
    last_exc: Exception | None = None
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=90) as resp:  # noqa: S310 (trusted gov hosts)
                return resp.read()
        except (urllib.error.URLError, TimeoutError) as exc:
            last_exc = exc
            logger.warning("GET %s failed (attempt %d/3): %s", url, attempt + 1, exc)
    raise RuntimeError(f"Failed to fetch {url}: {last_exc}")


def _sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def _notes_hash(text: str | None) -> str | None:
    if not text:
        return None
    return hashlib.sha1(text.encode("utf-8")).hexdigest()  # noqa: S324 (matches connector dedupe)


def resolve_realtime_urls() -> list[str]:
    try:
        raw = _http_get(DIRECTORY_ENDPOINT)
        payload = json.loads(raw.decode("utf-8"))
        data = payload.get("data", {}) if isinstance(payload, dict) else {}
        urls = [
            str(v)
            for k, v in data.items()
            if isinstance(v, str) and "realtime" in k.lower() and v.lower().endswith(".xlsx")
        ]
        if urls:
            return sorted(urls)
    except Exception as exc:  # pragma: no cover - network fallback
        logger.warning("Directory endpoint unavailable (%s); using fallback URLs.", exc)
    return list(FALLBACK_REALTIME_URLS)


def _parse_realtime_workbook(content: bytes, source_url: str) -> dict[tuple[int, int], FirstPrint]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        matrix = wb["cfnai_realtime"]
        rows = list(matrix.iter_rows(values_only=True))
        if not rows:
            return {}
        header = rows[0]
        colidx: dict[str, int] = {
            str(key): idx for idx, key in enumerate(header) if isinstance(key, str) and key
        }
        rowidx: dict[tuple[int, int], tuple[Any, ...]] = {}
        for row in rows[1:]:
            ref = row[0]
            if isinstance(ref, datetime):
                rowidx[(ref.year, ref.month)] = row

        release_by_month: dict[tuple[int, int], date] = {}
        for row in wb["release_dates"].iter_rows(values_only=True):
            release_dt, cfnai_month = row[0], row[1]
            if isinstance(release_dt, datetime) and isinstance(cfnai_month, datetime):
                release_by_month[(cfnai_month.year, cfnai_month.month)] = release_dt.date()
    finally:
        wb.close()

    out: dict[tuple[int, int], FirstPrint] = {}
    for (year, month), release_date in release_by_month.items():
        row = rowidx.get((year, month))
        if row is None:
            continue
        ck = colidx.get(f"CF{month}{year}")
        mk = colidx.get(f"CF3{month}{year}")
        if ck is None or mk is None:
            continue
        cval = row[ck]
        mval = row[mk]
        if not isinstance(cval, (int, float)) or not isinstance(mval, (int, float)):
            continue
        out[(year, month)] = FirstPrint(
            ref_year=year,
            ref_month=month,
            release_date=release_date,
            cfnai=float(cval),
            cfnai_ma3=float(mval),
            source_url=source_url,
        )
    return out


def collect_first_prints(url_payloads: list[tuple[str, bytes]]) -> dict[tuple[int, int], FirstPrint]:
    combined: dict[tuple[int, int], FirstPrint] = {}
    for source_url, content in url_payloads:
        parsed = _parse_realtime_workbook(content, source_url)
        for key, fp in parsed.items():
            existing = combined.get(key)
            # Keep the earliest release_date == the genuine first print if a
            # reference month is a headline edition in more than one workbook.
            if existing is None or fp.release_date < existing.release_date:
                combined[key] = fp
    return combined


def _month_key(value: str) -> tuple[int, int]:
    dt = datetime.strptime(value.strip(), "%Y-%m")
    return (dt.year, dt.month)


def _iter_month_keys(start: tuple[int, int], end: tuple[int, int]) -> list[tuple[int, int]]:
    if start > end:
        raise ValueError(f"CFNAI start month {start} is after end month {end}.")
    year, month = start
    out: list[tuple[int, int]] = []
    while (year, month) <= end:
        out.append((year, month))
        if month == 12:
            year, month = year + 1, 1
        else:
            month += 1
    return out


def _obs_period(year: int, month: int) -> str:
    return f"{year:04d}-{month:02d}-01"


def _build_records(
    first_prints: list[FirstPrint],
    *,
    cfnai_spec: MetricSpec,
    ma3_spec: MetricSpec,
    round_dp: int,
) -> tuple[list[ObservationRecord], list[ObservationRecord]]:
    retrieved = utc_now_iso()
    cfnai_notes = _notes_hash(cfnai_spec.notes)
    ma3_notes = _notes_hash(ma3_spec.notes)
    cfnai_records: list[ObservationRecord] = []
    ma3_records: list[ObservationRecord] = []
    for fp in first_prints:
        period = _obs_period(fp.ref_year, fp.ref_month)
        release = fp.release_date.isoformat()
        cfnai_records.append(
            _record(cfnai_spec, period, release, _round(fp.cfnai, round_dp), retrieved, cfnai_notes)
        )
        ma3_records.append(
            _record(ma3_spec, period, release, _round(fp.cfnai_ma3, round_dp), retrieved, ma3_notes)
        )
    return cfnai_records, ma3_records


def _round(value: float, dp: int) -> float:
    rounded = round(value, dp)
    return 0.0 if rounded == 0 else rounded


def _record(
    spec: MetricSpec,
    period: str,
    release: str,
    value: float,
    retrieved: str,
    notes_hash: str | None,
) -> ObservationRecord:
    return ObservationRecord(
        metric_key=spec.metric_key,
        source_name=spec.source_name,
        source_dataset=spec.source_dataset,
        source_series_id=spec.source_series_id,
        ref_area=spec.ref_area,
        frequency=spec.frequency,
        seasonal_adjustment=spec.seasonal_adjustment,
        units=spec.units,
        observation_period=period,
        observation_date=period,
        release_date=release,
        vintage_date=release,
        value=value,
        source_last_updated=None,
        retrieved_at=retrieved,
        revision_flag=0,
        notes_hash=notes_hash,
    )


def _spec_by_key(specs: list[MetricSpec], registry_key: str) -> MetricSpec:
    for spec in specs:
        if spec.registry_key == registry_key:
            return spec
    raise ValueError(f"Registry key {registry_key!r} not found in registry CSV.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=None, help="Path to macro raw YAML config.")
    parser.add_argument("--db-path", type=Path, default=None, help="Optional SQLite DB path override.")
    parser.add_argument("--start", type=str, default=DEFAULT_START_MONTH, help="Start reference month YYYY-MM (inclusive).")
    parser.add_argument("--end", type=str, default=DEFAULT_END_MONTH, help="End reference month YYYY-MM (inclusive).")
    parser.add_argument("--round-dp", type=int, default=VALUE_ROUND_DP, help="Decimal places (CFNAI is published to 2).")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print only (default behavior).")
    parser.add_argument("--apply", action="store_true", help="Write parsed rows to macro_observation_raw.")
    return parser.parse_args()


def main() -> None:
    configure_pipeline_logging()
    args = parse_args()
    apply = bool(args.apply)
    config_path = resolve_config_path(args.config)
    _, cfg = load_macro_raw_config(config_path)
    db_path = resolve_db_path(cfg, config_path, override=args.db_path)
    registry_csv = resolve_path(config_path, str(cfg.get("registry_csv") or "MacroLayer/macro_metric_registry_full.csv"))
    if registry_csv is None:
        raise ValueError("Could not resolve registry_csv from config.")
    specs = load_metric_registry(registry_csv)
    cfnai_spec = _spec_by_key(specs, CFNAI_REGISTRY_KEY)
    ma3_spec = _spec_by_key(specs, CFNAI_MA3_REGISTRY_KEY)

    start_key = _month_key(args.start)
    end_key = _month_key(args.end)

    urls = resolve_realtime_urls()
    url_payloads: list[tuple[str, bytes]] = []
    payload_shas: dict[str, str] = {}
    for url in urls:
        content = _http_get(url)
        url_payloads.append((url, content))
        payload_shas[url] = _sha256(content)
        logger.info("Fetched %s (%d bytes, sha256=%s)", url, len(content), payload_shas[url][:12])

    all_prints = collect_first_prints(url_payloads)
    expected_keys = _iter_month_keys(start_key, end_key)
    missing_keys = [key for key in expected_keys if key not in all_prints]
    if missing_keys:
        sample = ", ".join(f"{year:04d}-{month:02d}" for year, month in missing_keys[:12])
        suffix = "..." if len(missing_keys) > 12 else ""
        raise RuntimeError(
            f"CFNAI backfill is incomplete: {len(missing_keys)} requested month(s) are unrecoverable "
            f"from the downloaded workbooks ({sample}{suffix})."
        )
    selected = [all_prints[key] for key in expected_keys]

    cfnai_records, ma3_records = _build_records(
        selected, cfnai_spec=cfnai_spec, ma3_spec=ma3_spec, round_dp=args.round_dp
    )

    logger.info(
        "CFNAI first-print backfill: window %s..%s -> %d reference months, %d rows (%d CFNAI + %d MA3).",
        args.start,
        args.end,
        len(selected),
        len(cfnai_records) + len(ma3_records),
        len(cfnai_records),
        len(ma3_records),
    )
    for fp in selected:
        print(
            f"  {fp.ref_year:04d}-{fp.ref_month:02d}  release={fp.release_date.isoformat()}  "
            f"CFNAI={_round(fp.cfnai, args.round_dp):+.2f}  MA3={_round(fp.cfnai_ma3, args.round_dp):+.2f}"
        )

    manifest = {
        "script": "backfill_cfnai_first_prints.py",
        "generated_at_utc": utc_now_iso(),
        "mode": "apply" if apply else "dry_run",
        "window": {"start_month": args.start, "end_month": args.end},
        "reference_months": len(selected),
        "rows_parsed": len(cfnai_records) + len(ma3_records),
        "rows_cfnai": len(cfnai_records),
        "rows_cfnai_ma3": len(ma3_records),
        "round_dp": args.round_dp,
        "source_urls": urls,
        "payload_sha256": payload_shas,
        "db_path": str(db_path),
        "registry_keys": [CFNAI_REGISTRY_KEY, CFNAI_MA3_REGISTRY_KEY],
    }

    rows_written = 0
    if apply:
        rows_written = _apply(
            db_path=db_path,
            cfnai_records=cfnai_records,
            ma3_records=ma3_records,
            payload_shas=payload_shas,
        )
        manifest["rows_written"] = rows_written
        logger.info("Applied %d CFNAI first-print rows to %s", rows_written, db_path)
    else:
        logger.info("DRY RUN: no rows written. Re-run with --apply to persist.")

    manifest_path = _write_manifest(config_path, manifest)
    logger.info("Wrote manifest: %s", manifest_path)


def _apply(
    *,
    db_path: Path,
    cfnai_records: list[ObservationRecord],
    ma3_records: list[ObservationRecord],
    payload_shas: dict[str, str],
) -> int:
    conn = connect_sqlite(db_path)
    run_id = "cfnai_firstprint_" + datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    try:
        init_db(conn)
        for registry_key in (CFNAI_REGISTRY_KEY, CFNAI_MA3_REGISTRY_KEY):
            exists = conn.execute(
                "SELECT 1 FROM macro_metric_registry WHERE registry_key = ?", (registry_key,)
            ).fetchone()
            if not exists:
                raise ValueError(
                    f"Registry row {registry_key!r} is missing; run the macro raw pipeline first."
                )
        start_run(
            conn,
            run_id=run_id,
            mode="backfill",
            as_of_date=date.today().isoformat(),
            source_filter="fred_alfred:CFNAI,CFNAIMA3",
            dry_run=False,
            task_count=2,
            source_count=1,
        )
        artifacts = [
            SourceArtifact(
                registry_key=CFNAI_REGISTRY_KEY,
                source_name="chicagofed_cfnai_realtime",
                request_url=url,
                payload_hash=sha,
                http_status=200,
                fetched_at=utc_now_iso(),
                row_count=len(cfnai_records) + len(ma3_records),
                extra_json={"artifact_kind": "cfnai_realtime_xlsx"},
            )
            for url, sha in payload_shas.items()
        ]
        _insert_artifacts(conn, run_id=run_id, artifacts=artifacts)
        written = _upsert_observations(
            conn, run_id=run_id, registry_key=CFNAI_REGISTRY_KEY, observations=cfnai_records
        )
        written += _upsert_observations(
            conn, run_id=run_id, registry_key=CFNAI_MA3_REGISTRY_KEY, observations=ma3_records
        )
        finish_run(conn, run_id=run_id, status="completed", rows_written=written, error_count=0)
        conn.commit()
        return written
    except BaseException:
        conn.rollback()
        raise
    finally:
        conn.close()


def _write_manifest(config_path: Path, manifest: dict[str, Any]) -> Path:
    out_dir = resolve_path(config_path, "MacroLayer/out/first_print_backfill")
    if out_dir is None:
        out_dir = Path("out/first_print_backfill")
    out_dir.mkdir(parents=True, exist_ok=True)
    manifest_path = out_dir / "cfnai_first_prints_manifest.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, sort_keys=True)
    return manifest_path


if __name__ == "__main__":
    main()
